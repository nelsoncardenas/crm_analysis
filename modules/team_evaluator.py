"""A module to evaluate a team of architects based on company-specific files.

    Typical usage example:
    from modules.team_evaluator import TeamEvaluator

    evaluator = TeamEvaluator(conf)
    evaluator.evaluate()
"""
from typing import Tuple, List, Set

from openpyxl import load_workbook, Workbook


class TeamEvaluator:
    """Class to evaluate a team of architects."""

    def __init__(self, conf: dict) -> None:
        """Initializes the instance.

        Args:
            conf (str): configuration dictionary.
        """
        self.path_input = conf["PATHS"]["INPUT"]
        self.path_team = conf["PATHS"]["TEAM"]
        self.path_output = conf["PATHS"]["OUTPUT"]

        self.text_executed = conf["TEXT_EXECUTED"]
        self.text_programmed = conf["TEXT_PROGRAMMED"]
        self.state_map = conf["STATE_MAP"]

        self.col_team = conf["COL_TEAM"]
        self.col_execution = conf["COL_EXECUTION"]
        self.col_name = conf["COL_NAME"]
        self.col_state = conf["COL_STATE"]

        self.team_architect_list = list()

        self.book_input = load_workbook(self.path_input, data_only=True)
        self.book_team = load_workbook(self.path_team, data_only=True)
        self.book_output = Workbook()

        self.sheet_input = self.book_input["Hoja1"]
        self.sheet_equipo = self.book_team.active
        self.sheet_output = self.book_output.active

        self.errors_counter = dict()
        self.executed_counter = dict()

        self.team_architect_list, self.team_architect_set = self._load_team()

        self.architects_graded = set()

    def evaluate(self) -> None:
        """Evaluates the architects team."""
        print("---------- Empezando el análisis ----------\n")
        print(
            "Advertencia, recuerda que el nombre de la persona en el archivo debe coincidir totalmente con el nombre en el CRM"
        )
        self._error_analysis()
        self._evaluate_scheduled_and_posponed()
        self._evaluate_execution()
        self._save_output_sheet()

    def _load_team(self) -> Tuple[List, Set]:
        """Loads team member names to be analized."""
        print("Cargando el equipo de arquitectos.")
        team_architect_list = []
        row = 1
        while True:
            name = self.sheet_equipo.cell(row=row, column=self.col_team).value
            if name is None:
                break
            team_architect_list.append(name)
            row += 1
        team_architect_list.sort()
        team_architect_set = set(team_architect_list)
        print(f"El equipo de preventa tiene un tamaño de {len(team_architect_set)}.")
        return team_architect_list, team_architect_set

    def _error_analysis(self):
        """Fills counters for executions and errors."""
        print("Generando datos sobre errores.")
        row = 2
        cell_to_validate = self.sheet_input.cell(row=row, column=12).value
        while cell_to_validate is not None:
            execution = self.sheet_input.cell(row=row, column=self.col_execution).value
            execution = " ".join(execution.split()[:4])
            name = self.sheet_input.cell(row=row, column=self.col_name).value
            if execution == self.text_programmed and name in self.team_architect_set:
                state = self.state_map[
                    self.sheet_input.cell(row=row, column=self.col_state).value
                ]
                name_state = name + " + " + state
                self.errors_counter[name_state] = (
                    self.errors_counter.get(name_state, 0) + 1
                )
            if execution == self.text_executed and name in self.team_architect_set:
                self.executed_counter[name] = self.executed_counter.get(name, 0) + 1
            row += 1
            cell_to_validate = self.sheet_input.cell(row=row, column=12).value
        print(row - 1, "líneas analizadas.")

    def _evaluate_scheduled_and_posponed(self):
        """Evaluates scheduled and posponed related errors."""
        print("Evaluando propuestas programadas y aplazadas.")
        self.sheet_output.cell(row=1, column=1).value = "Nombre"
        self.sheet_output.cell(row=1, column=2).value = "Puntaje"
        self.sheet_output.cell(row=1, column=3).value = "Texto Errores"

        errores_keys = list(self.errors_counter.keys())
        errores_keys.sort()
        row = 0
        while row < len(errores_keys):
            architect_errors = 0
            name, _, num_actual_errors, errors_text = self._get_error_info(
                self.errors_counter, errores_keys[row]
            )
            architect_errors += num_actual_errors
            if (
                row < len(errores_keys) - 1
                and name == errores_keys[row + 1].split(" + ")[0]
            ):  # if the next position is the same architect
                name, _, num_actual_errors, errors_text = self._get_error_info(
                    self.errors_counter, errores_keys[row + 1], errors_text
                )
                architect_errors = architect_errors + num_actual_errors
                row += 1
            score = self._get_score(architect_errors)
            self.architects_graded.update([name])
            self.sheet_output.cell(
                row=self.team_architect_list.index(name) + 2, column=2
            ).value = score
            self.sheet_output.cell(
                row=self.team_architect_list.index(name) + 2, column=3
            ).value = errors_text
            row += 1

    def _get_error_info(
        self, errors_counter: dict, name_error: str, errors_text: str = ""
    ):
        """Generates error variables to perform analysis and report errors as text."""
        name, error = name_error.split(" + ")
        num_errores_actual = errors_counter[name_error]
        errors_text = f"{errors_text} Tiene {num_errores_actual} preventas {error} incorrectamente."  # noqa: E501
        return name, error, num_errores_actual, errors_text

    def _get_score(self, architect_errors: int) -> float:
        if architect_errors == 0:
            return 1.0
        if 1 <= architect_errors <= 2:
            return 0.5
        return 0.0

    def _evaluate_execution(self):
        """Evaluates proposals execution"""
        print("Evaluando ejecución de propuestas.")
        for name in self.team_architect_list:
            self.sheet_output.cell(
                row=self.team_architect_list.index(name) + 2, column=1
            ).value = name
            if name not in self.architects_graded:
                self.sheet_output.cell(
                    row=self.team_architect_list.index(name) + 2, column=2
                ).value = 1
            if name not in self.executed_counter:
                self.sheet_output.cell(
                    row=self.team_architect_list.index(name) + 2, column=2
                ).value = 0
                celda_errores = str(
                    self.sheet_output.cell(
                        row=self.team_architect_list.index(name) + 2, column=3
                    ).value
                )
                if celda_errores == "None":
                    self.sheet_output.cell(
                        row=self.team_architect_list.index(name) + 2, column=3
                    ).value = "No ha ejecutado preventas en las últimas dos semanas"
                else:
                    self.sheet_output.cell(
                        row=self.team_architect_list.index(name) + 2, column=3
                    ).value = (
                        celda_errores
                        + " No han ejecutado preventas en las últimas dos semanas."
                    )

    def _save_output_sheet(self):
        print(f"Guardando el resultado en {self.path_output}")
        self.book_output.save(self.path_output)
        print("\n¡Programa ejecutado correctamente!")
